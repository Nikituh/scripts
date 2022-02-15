
from openpyxl import load_workbook
from copy import copy
import types

workbook = load_workbook(filename = "file.xlsx")
worksheet = workbook.active

# Define row and column numbers where necessary data is stored
LEFT_COL_COUNT_NR  = 31
LEFT_COL_TRANSECT_NR  = 32
LEFT_COL_TRAIL_NR     = 33

RIGHT_COL_LISTING  = 35
RIGHT_COL_TRANSECT = 36

RIGHT_COL_COUNT_NR    = 38
RIGHT_COL_TRANSECT_NR = 39
RIGHT_COL_TRAIL_NR    = 40

START_ROW = 10
START_COL = 2
START_ROW_RIGHT = 36
MAX_SPECIES_COL = 28

# Variable to hold data collected from left table
data = []

# Loop over all the rows, til the last filled row
for row in range(START_ROW, worksheet.max_row):
    # Store row data in simple object
    result = types.SimpleNamespace()
    result.species_list = []
    
    # Loop over columns, starting from first species column til the max constant
    for column in range(START_COL, MAX_SPECIES_COL):
        species = worksheet[row][column].value
        if species == None or species == "" or species.isspace():
            # If the value for cell is 'None', it's empty. Break out of this loop
            break
        result.species_list.append(species.strip())
    
    # Save entire head row, so we could also copy style later
    result.head = worksheet[row][0]
    result.transect = worksheet[row][1].value

    result.count = len(result.species_list)
    result.transect_nr = worksheet[row][LEFT_COL_TRANSECT_NR].value
    result.trail_nr = worksheet[row][LEFT_COL_TRAIL_NR].value
    

    data.append(result)

# We start the loop from the first empty row in the right table, hardcoded
row_counter = START_ROW_RIGHT

# Now loop over the data we previously collected and fill the table on the right side
for item in data:
    header_cell = worksheet[row_counter][RIGHT_COL_LISTING]
    header_cell.value = item.head.value
    header_cell.fill = copy(item.head.fill)
    header_cell.border = copy(item.head.border)
    header_cell.font = copy(item.head.font)

    worksheet[row_counter][RIGHT_COL_TRANSECT].value = item.transect

    worksheet[row_counter][RIGHT_COL_COUNT_NR].value = item.count
    worksheet[row_counter][RIGHT_COL_TRANSECT_NR].value = item.transect_nr
    worksheet[row_counter][RIGHT_COL_TRAIL_NR].value = item.trail_nr

    for species in item.species_list:
        # Loop over species, increasing row number every time
        row_counter += 1
        worksheet[row_counter][RIGHT_COL_LISTING].value = species

    row_counter += 1

# Save output
OUTPUT_FILE = "draft.xlsx"
workbook.save(OUTPUT_FILE)

# The resuling xlsx is corrupted and needs to be "repaired" (it's actually fine)
# Just opening and saving it again was supposed to fix it, but doesn't do anything
workbook = load_workbook(filename = OUTPUT_FILE)
workbook.save(OUTPUT_FILE)


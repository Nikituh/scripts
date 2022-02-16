
from openpyxl import load_workbook
import types
import sys

FILENAME = "ruutude-arvutus-piiale-veebr2022"

workbook = load_workbook(filename = FILENAME + ".xlsx")
worksheet = workbook.active

# First gather all types into an object. This is a simple sheet, no need to define constants
type_sheet = workbook["elupaigatüüp_19"]
type_dict = {}
type_rows_total = type_sheet.max_row
for row in range(1, type_rows_total + 1):
	id = type_sheet[row][0].value
	type = type_sheet[row][1].value
	type_dict[id] = type

REPLACE_ID_START_ROW = 588
MAX_ROW = worksheet.max_row

PLANT_LISTING_COL_LEGACY = 0
PLANT_LISTING_COL = 1

for row in range(REPLACE_ID_START_ROW, MAX_ROW + 1):
	sys.stdout.write("\r")
	sys.stdout.write("Replacing legacy field name: " + str(row) + "/" + str(MAX_ROW))
	sys.stdout.flush()
	
	a_value = worksheet[row][PLANT_LISTING_COL_LEGACY].value
	if a_value is not None and a_value.startswith("2019"):
		worksheet[row][PLANT_LISTING_COL].value = a_value
		worksheet[row][PLANT_LISTING_COL_LEGACY].value = None

sys.stdout.write("\n")

BOX_CALCULATION_START_ROW = 43

# Collect all data from left columns and store in 'data' object
data = []
# Store row data in simple object
item = types.SimpleNamespace()
item.plants = []

initial = True
for row in range(BOX_CALCULATION_START_ROW, MAX_ROW + 1):

	sys.stdout.write("\r")
	sys.stdout.write("Processing row data: " + str(row) + "/" + str(MAX_ROW))
	sys.stdout.flush()

	heading_text = worksheet[row][PLANT_LISTING_COL].value

	if initial:
		item.id = worksheet[row][1].value
		item.row_nr = row
		initial = False

	if heading_text is not None and heading_text.startswith("20"):
		# If it starts with '20', it's a new object start row.
		# Add current iteration object to result and clear previous item data collected
		if len(item.plants) != 0:
			# Only add item if plants have been added, else there's no point. Happesn on inital row
			data.append(item)

		item = types.SimpleNamespace()
		item.plants = []
		item.id = worksheet[row][1].value
		item.row_nr = row
		if item.id.startswith("2019"):
			# Only 2019 environments exist in type sheet, and the only ones necessary as well
			item.environment = type_dict[item.id]
		
	elif heading_text is not None:
		plant = types.SimpleNamespace()
		plant.habitat             = worksheet[row][2].value
		plant.human_tolerance     = worksheet[row][3].value
		plant.frequency           = worksheet[row][4].value
		plant.border_proportion   = worksheet[row][5].value
		plant.conservation_status = worksheet[row][6].value
		plant.red_book            = worksheet[row][7].value
		plant.average_height      = worksheet[row][8].value
		plant.max_height          = worksheet[row][9].value
		plant.est_end             = worksheet[row][10].value
		plant.est_begin           = worksheet[row][11].value
		plant.est_persistence     = worksheet[row][12].value
		item.plants.append(plant)


ENVIRONMENT_COL     = 14
IDENTIFIER_COL      = 15
GRASSLAND_COL       = 16
FOREST_COL          = 17
WETLAND_COL         = 18
HUMAN_TOLERANCE_COL = 19
FREQUENCE_COL       = 20
PROPORTION_COL      = 21
CONSERV_STATUS_COL  = 22
RED_BOOK_COL        = 23
AVERAGE_HEIGHT_COL  = 24
MAX_HEIGHT_COL      = 25
EST_END_COL         = 26
EST_BEGIN_COL       = 27
EST_PERSISTENCE_COL = 28

def numeric(input):
	if input == None:
		return 0

	if isinstance(input, str):

		stripped = input.strip()

		if stripped == "hemerofoob":
			return 1
		if stripped == "hemeradiafoor":
			return 2
		if stripped == "apofüüt" or stripped == "antropofüüt":
			return 3

		if stripped == "paiguti" or stripped == "hajusalt" or stripped == "haruldane":
			return 1
		if stripped == "tavaline":
			return 2
		if stripped == "sage":
			return 3

	return float(input)

for item in data:
	
	row = worksheet[item.row_nr]

	grassland_total           = 0
	forest_total              = 0
	wetland_total             = 0

	human_tolerance_total     = 0
	frequency_total           = 0
	border_proportion_total   = 0
	conservation_status_total = 0
	red_book_total            = 0
	average_height_total      = 0
	max_height_total          = 0

	est_end_total             = 0
	est_begin_total           = 0
	est_persistence_total     = 0
	
	for plant in item.plants:


		if plant.habitat == "niit":
			grassland_total += 1
		elif plant.habitat == "mets":
			forest_total += 1
		elif plant.habitat == "märgala":
			wetland_total += 1

		human_tolerance_total     += numeric(plant.human_tolerance)
		frequency_total           += numeric(plant.frequency)
		
		if plant.border_proportion == "yes":
			border_proportion_total += 1
		if plant.conservation_status == "yes":
			conservation_status_total += 1
		if plant.red_book == "yes":
			red_book_total += 1
		
		average_height_total      += numeric(plant.average_height)
		max_height_total          += numeric(plant.max_height)

		est_end_total             += numeric(plant.est_end)
		est_begin_total           += numeric(plant.est_begin)
		est_persistence_total     += numeric(plant.est_persistence)

	count = len(item.plants)

	if item.id.startswith("2019"):
		row[ENVIRONMENT_COL].value = item.environment

	row[IDENTIFIER_COL].value = item.id
	
	row[GRASSLAND_COL].value = grassland_total / count
	row[FOREST_COL].value = forest_total / count
	row[WETLAND_COL].value = wetland_total / count
	
	row[HUMAN_TOLERANCE_COL].value = human_tolerance_total / count
	row[FREQUENCE_COL].value = frequency_total / count
	row[PROPORTION_COL].value = border_proportion_total / count
	row[CONSERV_STATUS_COL].value = conservation_status_total / count
	row[RED_BOOK_COL].value = red_book_total / count
	row[AVERAGE_HEIGHT_COL].value = average_height_total / count
	row[MAX_HEIGHT_COL].value = max_height_total / count
	row[EST_END_COL].value = est_end_total / count
	row[EST_BEGIN_COL].value = est_begin_total / count
	row[EST_PERSISTENCE_COL].value = est_persistence_total / count

workbook.save(FILENAME + "-output.xlsx")

print("Done! Processed " + str(MAX_ROW) + " rows")





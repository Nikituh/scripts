
# Compile using pyinstaller: pyinstaller edl.py

import csv
from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border

#### CONSTANTS ####

DURATION_LINE_IDENTIFIER = "  AX       "
AUTHOR_LINE_IDENTIFIER = "* FROM CLIP NAME:"

ARTIST_IDENTIFIER = "(A)"
PRESENTER_IDENTIFIER = "(PRS)"
# Just another tag: The American Society of Composers, Authors and Publishers
ASCAP_IDENTIFIER = "(ASCAP)"
# Broadcast Music, Inc. (BMI) is one of three United States performing rights organizations, along with ASCAP and SESAC.
BMI_IDENTIFIER = "(BMI)"

WOM_IDENTIFIER = "WOM_"
TSH_IDENTIFIER = "TSH"
RFM_IDENTIFIER = "RFM"

WAV_IDENTIFIER = ".wav"
MP3_IDENTIFIER = ".mp3"

NEWLINE = "\n"
UNDERSCORE = "_"
COLON = ":"
SPACE = " "
DASH = "-"

HOUR_CONSTANT = 60 * 60
MINUTE_CONSTANT = 60
		
#### ENTRY CLASS ####

class Entry(object):
	
	artist = ""
	song = ""
	label = ""
	
	hours = 0
	minutes = 0
	seconds = 0

	def __init__(self):
		super(Entry, self).__init__()

	def set_artist_from_wom(self, string):
		self.set_data_from_label(string)

	def set_artist_from_tsh(self, string):
		self.set_data_from_label(string)

	def set_artist_from_rfm(self, string):
		self.set_data_from_label(string)

	def set_artist_from_other(self, string):
		
		split = string.split(DASH)

		if len(split) == 2:
			self.artist = split[0].strip()
			self.song = split[1].strip().replace(WAV_IDENTIFIER, "").replace(MP3_IDENTIFIER, "")
		else:
			# If there's more than one dash, we have no idea what's the song or artist,
			# just add the entire line so it can later be manually fixed
			self.artist = string.replace(WAV_IDENTIFIER, "").replace(MP3_IDENTIFIER, "")

		self.label = "NONE"

	def set_duration(self, duration):

		split = duration.split(":")
		
		self.hours += int(split[0])
		self.minutes += int(split[1])
		self.seconds += int(split[2])

	def get_duration_in_seconds(self):

		total_seconds = (self.hours * 60 * 60) + (self.minutes * 60) + self.seconds

		# hours
		hours = total_seconds // 3600 
		
		# remaining seconds
		total_seconds = total_seconds - (hours * 3600)
		
		# minutes
		minutes = total_seconds // 60
		# remaining
		seconds = total_seconds - (minutes * 60)
		
		if (hours < 9):
			hours = "0" + str(hours)
		
		if (minutes < 9):
			minutes = "0" + str(minutes)

		if (seconds < 9):
			seconds = "0" + str(seconds)

		return '%s:%s:%s' % (hours, minutes, seconds)

	def set_data_from_label(self, string):

		string = self.remove_label_identifier(string)

		splitByArtistIdentifier = string.rpartition(ARTIST_IDENTIFIER)
		
		song = splitByArtistIdentifier[0]
		song = song.replace(UNDERSCORE, SPACE)
		self.song = song.strip().title()

		remainder = splitByArtistIdentifier[2]

		if (PRESENTER_IDENTIFIER in remainder):
			partitioner = PRESENTER_IDENTIFIER
		elif (ASCAP_IDENTIFIER in remainder):
			partitioner = ASCAP_IDENTIFIER
		elif BMI_IDENTIFIER in remainder:
			partitioner = BMI_IDENTIFIER

		remainder = remainder.rpartition(partitioner);

		artist = remainder[0]

		artist = artist.replace(UNDERSCORE, SPACE)
		self.artist = artist.strip().title()
		
		remainder = remainder[2]
		remainder = remainder.replace(UNDERSCORE, SPACE).replace(WAV_IDENTIFIER, "").replace(MP3_IDENTIFIER, "")

		self.label = remainder.strip();

	def remove_label_identifier(self, string):
		
		# The label_identifier will be along the lines of "WOM_405_trk008_"
		# Important part: song starts after the third underscore
		split = string.split(UNDERSCORE)
		
		result = string.replace(split[0], "").replace(split[1], "").replace(split[2], "")
		result = result.strip(UNDERSCORE)

		return result;

	def is_song_of_artist(self, entry):
		return self.song == entry.song and self.artist == entry.artist

#### PROGRAM START ####

filename = raw_input("Enter the name of the .edl you'd like the parse (e.g. 8.edl): ")

lines = [line for line in open(filename)]

total = len(lines)

entries = []

for i in range(0, total - 1):
	line = lines[i]
	
	if DURATION_LINE_IDENTIFIER in line:
		
		duration_line = line;
		artist_line = lines[i + 1]

		artist_line = artist_line.replace("* FROM CLIP NAME:", "").strip()

		if (WAV_IDENTIFIER not in artist_line and MP3_IDENTIFIER not in artist_line):
			# If it's not an audio file, ignore, move on
			continue

		split = duration_line.split(" ")
		duration = split[len(split) - 3]

		entry = Entry()
		entry.set_duration(duration)
		
		if (artist_line.startswith(WOM_IDENTIFIER)):
			entry.set_artist_from_wom(artist_line)
		elif (artist_line.startswith(TSH_IDENTIFIER)):
			entry.set_artist_from_tsh(artist_line)
		elif (artist_line.startswith(RFM_IDENTIFIER)):
			entry.set_artist_from_rfm(artist_line)
		else:
			entry.set_artist_from_other(artist_line)

		exists = False

		# Add seconds if artist & song already exist
		for existing in entries:

			if existing.is_song_of_artist(entry):
				existing.set_duration(duration)
				exists = True

		if (exists):
			continue

		entries.append(entry)

#### FILE WRITE ####

wb = Workbook()

ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = "Artist"
ws["B1"] = "Song"
ws["C1"] = "Duration"
ws["D1"] = "Label"

light_green = PatternFill(start_color='7CFC00', end_color='7CFC00', fill_type='solid')

ws['A1'].fill = light_green
ws['B1'].fill = light_green
ws['C1'].fill = light_green
ws['D1'].fill = light_green

for entry in entries:
	ws.append([entry.artist, entry.song, entry.get_duration_in_seconds(), entry.label])

wb.save(filename + ".xlsx")



